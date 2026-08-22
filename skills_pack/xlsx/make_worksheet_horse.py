import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

ARTICLE_LINES = [
    '有個商人養了一匹馬和一頭驢子。',
    '有次出門前，他先把貨物放在驢子身上，直到驢子受不了，才把剩下的一點點貨物放在馬身上。',
    '路途中，驢子因為身體很不舒服，就對馬說：「馬大哥，我的負擔實在太重，請幫忙分擔一點，好嗎？我已經快走不動了！」馬搖搖頭，拒絕了驢子的請求。',
    '後來，驢子過於勞累，就倒下死了。於是，主人把驢子背上所有的貨物，以及那張驢子皮，全都放在馬背上。',
    '這時，馬很後悔，心裡想：「這都是因為不願意幫驢子忙的結果。現在，我不但要背上全部的貨物，還多加了一張驢皮。真是      ！」',
]

thin = Side(style='thin', color='AAAAAA')
border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
border_bottom = Border(bottom=Side(style='thin', color='888888'))

def cell_set(ws, r, value, font=None, fill=None, align=None, border=None, height=None):
    ws.merge_cells(f'A{r}:B{r}')
    c = ws[f'A{r}']
    c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = border
    if height: ws.row_dimensions[r].height = height
    return r + 1

def make_sheet(wb, name, color_hex, title, sub, level, students, design_note, qlist, alist):
    ws = wb.create_sheet(name)
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 78
    r = 1

    # 標題
    r = cell_set(ws, r, title,
        Font(name='標楷體', size=16, bold=True, color='FFFFFF'),
        PatternFill('solid', fgColor=color_hex),
        Alignment(horizontal='center', vertical='center', wrap_text=True),
        height=32)

    # 副標
    r = cell_set(ws, r, sub,
        Font(name='標楷體', size=11, bold=True),
        PatternFill('solid', fgColor='F0F0F0'),
        Alignment(horizontal='left', vertical='center'),
        height=20)

    # 層次說明
    r = cell_set(ws, r, f'◆ 閱讀層次：{level}　　◆ 對應學生：{students}',
        Font(name='標楷體', size=10, color='444444'),
        PatternFill('solid', fgColor='FAFAFA'),
        Alignment(horizontal='left', vertical='center'),
        height=18)

    # 降階設計
    r = cell_set(ws, r, f'◆ 降階設計：{design_note}',
        Font(name='標楷體', size=10, color='444444'),
        PatternFill('solid', fgColor='FAFAFA'),
        Alignment(horizontal='left', vertical='center'),
        height=18)

    # 色條分隔
    ws.merge_cells(f'A{r}:B{r}')
    ws[f'A{r}'].fill = PatternFill('solid', fgColor=color_hex)
    ws.row_dimensions[r].height = 4
    r += 1

    # 文章標題
    r = cell_set(ws, r, '【閱讀文章】馬與驢子（113年三年級學習扶助篩選測驗）',
        Font(name='標楷體', size=12, bold=True),
        PatternFill('solid', fgColor='E8F4FD'),
        Alignment(horizontal='center', vertical='center'),
        height=22)

    # 文章內容
    for line in ARTICLE_LINES:
        r = cell_set(ws, r, line,
            Font(name='標楷體', size=12),
            PatternFill('solid', fgColor='EEF7FF'),
            Alignment(horizontal='left', vertical='center', wrap_text=True),
            height=24)

    # 色條分隔
    ws.merge_cells(f'A{r}:B{r}')
    ws[f'A{r}'].fill = PatternFill('solid', fgColor=color_hex)
    ws.row_dimensions[r].height = 4
    r += 1

    # 題目
    for qi, (q, ans) in enumerate(zip(qlist, alist), 1):
        r = cell_set(ws, r, f'第{qi}題',
            Font(name='標楷體', size=11, bold=True, color='FFFFFF'),
            PatternFill('solid', fgColor=color_hex),
            Alignment(horizontal='left', vertical='center'),
            border_all, 20)

        for line in q.split('\n'):
            r = cell_set(ws, r, line,
                Font(name='標楷體', size=12),
                None,
                Alignment(horizontal='left', vertical='center', wrap_text=True),
                border_all, 22)

        for _ in range(3):
            ws.merge_cells(f'A{r}:B{r}')
            c = ws[f'A{r}']
            c.fill = PatternFill('solid', fgColor='FFFEF0')
            c.border = border_bottom
            ws.row_dimensions[r].height = 22
            r += 1

        ws.row_dimensions[r].height = 6
        r += 1

    # 教師參考答案
    r = cell_set(ws, r, '【教師參考答案】',
        Font(name='標楷體', size=11, bold=True, color='FFFFFF'),
        PatternFill('solid', fgColor='888888'),
        Alignment(horizontal='center', vertical='center'),
        height=22)

    for qi, ans in enumerate(alist, 1):
        r = cell_set(ws, r, f'第{qi}題：{ans}',
            Font(name='標楷體', size=10, color='333333'),
            PatternFill('solid', fgColor='F5F5F5'),
            Alignment(horizontal='left', vertical='center', wrap_text=True),
            height=24)

    return ws

# ===== A卷 =====
qA = [
    '根據文章，驢子向馬說了什麼？請在文章中圈出來，再把答案抄在下面的橫線上。\n（提示：在文章第三段找找看。）\n\n答：',
    '驢子倒下死了之後，主人把哪些東西放在馬背上？請在文章第四段找一找，圈出後填在橫線上。\n\n答：',
    '故事最後，馬有什麼感受？請用文章裡的一個詞語回答。\n\n答：',
]
ansA = [
    '「馬大哥，我的負擔實在太重，請幫忙分擔一點，好嗎？我已經快走不動了！」',
    '驢子背上所有的貨物，以及那張驢子皮',
    '後悔',
]
make_sheet(wb, 'A卷（提取訊息）', '4472C4',
    'A 卷（第一組：學生G、H）── 提取訊息題',
    '學習目標：能找出文章中明確寫出的訊息，答案可直接在文章中找到。',
    '①提取訊息（定位型，無需推論）',
    '第一組（學生G、H，重度補救）',
    '直接告訴學生去哪一段找，降低搜尋負擔',
    qA, ansA)

# ===== B卷 =====
qB = [
    '根據文章，驢子倒下死掉的原因是什麼？請選出最正確的答案，並完成下面的推論句。\n\n(1) 牠被馬欺負而病死\n(2) 馬不願幫牠而氣死\n(3) 商人把馬背的貨物放在牠背上\n(4) 牠背非常重的貨物而過度疲累\n\n我選（　　）。\n因為文章說：「                                  」，\n所以我推論驢子是因為                而倒下死了。',
    '文章最後，馬心裡說「真是　　　　！」哪個詞語最適合填在空格裡？\n\n(1) 自作自受　(2) 自動自發　(3) 自言自語　(4) 自由自在\n\n我選（　　）。\n理由：馬                ，所以現在                ，\n這就是「    　　」的意思。',
    '如果馬當初願意幫驢子分擔貨物，結果會怎麼不一樣？請根據文章推論，完成下面的句子。\n\n如果馬當初願意幫驢子，驢子就不會                ，\n馬也不需要                              。',
]
ansB = [
    '選(4)。文章說：「驢子過於勞累，就倒下死了」，所以推論驢子是因為背了非常重的貨物、過度疲累而倒下死了。',
    '選(1)自作自受。馬不願幫忙，結果驢子死了，主人把所有貨物加驢皮都放到馬背上，這就是「自作自受」（自己做錯事，自己承擔後果）。',
    '驢子就不會因為過度勞累而倒下死亡；馬也不需要背上全部貨物加上一張驢皮。',
]
make_sheet(wb, 'B卷（推論訊息）', '70AD47',
    'B 卷（第二組：學生D、E）── 推論訊息題',
    '學習目標：能根據文章的線索，推論出文章沒有直說的原因或結果。',
    '②推論訊息（因果推論、詞義推論）',
    '第二組（學生D、E，中度補救）',
    '含提示框引導「文章說…，所以推論…」，降低推論負擔',
    qB, ansB)

# ===== C卷 =====
qC = [
    '這篇文章主要想告訴我們什麼道理？請選出最適合的答案，並用文章中至少一句話說明你的理由。\n\n(1) 體力充沛比頭腦聰明更重要\n(2) 有能力的人應該分擔更多工作\n(3) 團隊合作對他人和自己都有好處\n(4) 不要只想賺錢而不在乎動物的性命\n\n我選（　　）。\n文章中有一段話說：「                                          」，\n可以看出                                ，\n所以這篇文章想告訴我們：                。',
    '有人說這篇文章的主題是「自私的代價」，也有人說是「互助合作的重要性」。\n你比較同意哪個說法？請用文章的內容支持你的看法。\n\n我比較同意「              」這個說法，\n因為文章中                                          ，\n所以我認為                              。',
    '讀完這個故事，對你有什麼啟示？請舉一個生活中可以互相幫忙的例子，說明為什麼幫助別人對自己也有好處。\n\n我的啟示：                                          。\n生活例子：                                          。',
]
ansC = [
    '選(3)。文章中「馬很後悔，心裡想：這都是因為不願意幫驢子忙的結果……真是自作自受！」可以看出不合作導致自己付出更大代價，說明合作對彼此都有好處。',
    '開放式，兩個說法均可接受，須引用文章具體內容支持觀點，言之有據即可。',
    '開放式，能連結文章主旨並舉出具體生活例子即可。',
]
make_sheet(wb, 'C卷（詮釋整合）', 'C55A11',
    'C 卷（第三組：學生F）── 詮釋整合題',
    '學習目標：能整合全文，說出作者要傳達的核心道理，並用文章例子支持自己的看法。',
    '③詮釋整合（主旨判讀、觀點論述）',
    '第三組（學生F，輕度補救）',
    '完整版考題，含「選→引文→說明」三段論述鷹架，附自我評估欄',
    qC, ansC)

del wb['Sheet']

out = 'D:/備課ai/研習講義/激發潛能之星/新北市五股區更寮國民小學_吳國榮老師/附件2-4/其他附件資料/04_差異化學習單_馬與驢子.xlsx'
wb.save(out)
print('完成！儲存至:', out)
